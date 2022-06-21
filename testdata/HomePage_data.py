import openpyxl


class HomePageData:
    # test_data = [
    #     {"Name": "Nishith Patnaik", "Email": "nishith@gmail.com", "Password": "ohno-ohno", "Gender": "Male"},
    #     {"Name": "Cristiano Ronaldo", "Email": "CR7@gmail.com", "Password": "ohyes-ohyes", "Gender": "Female"}]

    # Parameterize using Excel Sheet
    # test_data = []

    @staticmethod  # WHEN A METHOD IS STATIC, it can be called using CLASSNAME.METHODNAME
    # NO NEED TO CREATE OBJECT of the class to call its method.
    # SELF PARAMETER IS NOT REQUIRED WHEN STATIC
    def get_excel_data():
        book = openpyxl.load_workbook(
            "C:\\Users\\CR7\\PycharmProjects\\PythonFrameworkProject\\testdata\\excelTestData.xlsm")
        #  Controls the sheet that was last active
        sheet = book.active
        lst = []
        test_data = []
        # returns each row from an Excel sheet as a dictionary stored within a list
        # Traversing Excel Sheet
        for r in range(2, sheet.max_row + 1):
            test_data = test_data + lst
            lst = []
            dict = {}
            for c in range(1, sheet.max_column + 1):
                # dict[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value
                # Row 1 is the column heading
                dict[sheet.cell(row=1, column=c).value] = sheet.cell(row=r,
                                                                     column=c).value  # Row 1 is the column heading
            lst.append(dict)

        return test_data
